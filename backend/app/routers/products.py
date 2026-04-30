import csv
import io
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import List, Optional
from app.database import get_db
from app.models.inventory import Product
from app.schemas import ProductCreate, ProductOut, ProductUpdate, StockAdjust, ProductImportResult
from app.core.deps import require_any, require_supervisor
from app.models.user import User

router = APIRouter(prefix="/products", tags=["products"])


def _normalize_header(value: str) -> str:
    return (value or "").strip().lower().replace(" ", "_")


def _as_text(row: dict, *keys: str, default: str = "") -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return default


def _as_float(row: dict, *keys: str, default: float = 0.0) -> float:
    raw = _as_text(row, *keys, default="")
    if raw == "":
        return default
    cleaned = raw.replace("$", "").replace(",", "").strip()
    try:
        return float(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        raise ValueError(f"Valor numérico inválido: {raw}")


def _as_int(row: dict, *keys: str, default: int = 0) -> int:
    raw = _as_text(row, *keys, default="")
    if raw == "":
        return default
    try:
        return int(float(raw))
    except ValueError:
        raise ValueError(f"Valor entero inválido: {raw}")


def _load_rows(filename: str, content: bytes) -> list[dict]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        text_content = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_content))
        return [{_normalize_header(k): v for k, v in row.items()} for row in reader]

    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(str(value or "")) for value in rows[0]]
        parsed_rows: list[dict] = []
        for values in rows[1:]:
            parsed_rows.append({
                headers[index]: values[index] if index < len(values) else None
                for index in range(len(headers))
                if headers[index]
            })
        return parsed_rows

    raise HTTPException(status_code=400, detail="Formato no soportado. Usa .csv o .xlsx")


@router.get("", response_model=List[ProductOut])
async def list_products(
    category: Optional[str] = None,
    active_only: bool = True,
    in_stock: bool = False,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_any),
):
    q = select(Product)
    if active_only:
        q = q.where(Product.is_active == True)
    if in_stock:
        q = q.where(Product.stock > 0)
    if category:
        q = q.where(Product.category.ilike(f"%{category}%"))
    q = q.order_by(Product.ai_priority.desc(), Product.name).offset(skip).limit(limit)
    result = await db.execute(q)
    return [ProductOut.model_validate(p) for p in result.scalars().all()]


@router.post("", response_model=ProductOut, status_code=201)
async def create_product(
    body: ProductCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_supervisor),
):
    product = Product(**body.model_dump())
    db.add(product)
    await db.commit()
    await db.refresh(product)
    return ProductOut.model_validate(product)


@router.get("/{product_id}", response_model=ProductOut)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_any),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return ProductOut.model_validate(p)


@router.patch("/{product_id}", response_model=ProductOut)
async def update_product(
    product_id: int,
    body: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_supervisor),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(p, field, value)
    await db.commit()
    await db.refresh(p)
    return ProductOut.model_validate(p)


@router.post("/{product_id}/stock", response_model=ProductOut)
async def adjust_stock(
    product_id: int,
    body: StockAdjust,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_supervisor),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    new_stock = p.stock + body.delta
    if new_stock < 0:
        raise HTTPException(status_code=400, detail="Stock no puede ser negativo")
    p.stock = new_stock
    await db.commit()
    await db.refresh(p)
    return ProductOut.model_validate(p)


@router.post("/import", response_model=ProductImportResult)
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_supervisor),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo inválido")

    content = await file.read()
    rows = _load_rows(file.filename, content)
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene filas para importar")

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    for index, row in enumerate(rows, start=2):
        try:
            name = _as_text(row, "name", "nombre")
            category = _as_text(row, "category", "categoria")
            if not name or not category:
                skipped += 1
                errors.append(f"Fila {index}: requiere nombre y categoría")
                continue

            sku = _as_text(row, "sku", default="")
            price = _as_float(row, "price", "precio")
            cost = _as_float(row, "cost", "costo", default=0.0)
            stock = _as_int(row, "stock", "inventario", "existencia", default=0)
            margin_pct = _as_float(row, "margin_pct", "margen", default=0.0)
            ai_priority = _as_int(row, "ai_priority", "prioridad_ia", default=5)
            description = _as_text(row, "description", "descripcion", default="")
            image_url = _as_text(row, "image_url", "imagen", default="") or None
            is_active_text = _as_text(row, "is_active", "activo", default="true").lower()
            is_active = is_active_text not in {"0", "false", "no", "inactivo"}

            existing = None
            if sku:
                existing = (
                    await db.execute(select(Product).where(Product.sku == sku))
                ).scalar_one_or_none()

            if not existing:
                existing = (
                    await db.execute(select(Product).where(func.lower(Product.name) == name.lower()))
                ).scalar_one_or_none()

            if existing:
                existing.category = category
                existing.description = description or existing.description
                existing.price = price
                existing.cost = cost
                existing.stock = stock
                existing.margin_pct = margin_pct
                existing.ai_priority = min(10, max(1, ai_priority))
                existing.image_url = image_url
                existing.is_active = is_active
                if sku:
                    existing.sku = sku
                updated += 1
            else:
                db.add(Product(
                    name=name,
                    category=category,
                    description=description or None,
                    price=price,
                    cost=cost,
                    stock=stock,
                    margin_pct=margin_pct,
                    sku=sku or None,
                    image_url=image_url,
                    ai_priority=min(10, max(1, ai_priority)),
                    is_active=is_active,
                ))
                created += 1
        except ValueError as exc:
            skipped += 1
            errors.append(f"Fila {index}: {exc}")

    await db.commit()
    return ProductImportResult(created=created, updated=updated, skipped=skipped, errors=errors[:20])
